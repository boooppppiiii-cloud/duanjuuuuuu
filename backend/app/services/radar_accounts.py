from __future__ import annotations

import csv
import io
import re
from datetime import date, datetime
from typing import Any
from urllib.parse import parse_qsl, urlencode, urlsplit, urlunsplit

from openpyxl import Workbook, load_workbook
from sqlmodel import Session, select

from ..models import (
    Drama,
    DramaSearchProfile,
    MonitoredAccount,
    MonitoredAccountDrama,
    RadarImportBatch,
    RadarUnmatchedDrama,
)


FIELDS = [
    "platform", "display_name", "platform_account_id", "profile_url",
    "relationship_type", "authorization_status", "company_name", "dramas",
    "authorized_regions", "authorization_start", "authorization_end",
    "allowed_full_series", "notes",
]

FIELD_LABELS = {
    "platform": "平台", "display_name": "账号名称", "platform_account_id": "账号ID",
    "profile_url": "账号主页", "relationship_type": "账号类型",
    "authorization_status": "授权状态", "company_name": "所属公司",
    "dramas": "负责剧目", "authorized_regions": "授权地区",
    "authorization_start": "授权开始时间", "authorization_end": "授权结束时间",
    "allowed_full_series": "是否允许完整剧", "notes": "备注",
}

HEADER_ALIASES = {
    "platform": {"平台", "platform", "渠道", "social platform"},
    "display_name": {"账号名称", "账号名", "display name", "account name", "channel name", "name"},
    "platform_account_id": {"账号id", "频道id", "account id", "channel id", "platform account id"},
    "profile_url": {"账号主页", "主页", "账号链接", "频道链接", "profile url", "channel url", "url"},
    "relationship_type": {"账号类型", "关系类型", "account type", "relationship type"},
    "authorization_status": {"授权状态", "authorization status", "auth status"},
    "company_name": {"所属公司", "公司", "company", "company name"},
    "dramas": {"负责剧目", "剧目", "剧名", "dramas", "drama", "series"},
    "authorized_regions": {"授权地区", "地区", "regions", "authorized regions", "region"},
    "authorization_start": {"授权开始时间", "授权开始", "start date", "authorization start"},
    "authorization_end": {"授权结束时间", "授权结束", "end date", "authorization end"},
    "allowed_full_series": {"是否允许完整剧", "允许完整剧", "allow full series", "allowed full series"},
    "notes": {"备注", "notes", "note"},
}

RELATIONSHIPS = {"own_official", "own_creator", "authorized_partner", "known_distributor", "unknown", "blocked"}
AUTH_STATUSES = {"authorized", "partial", "expired", "unknown", "unauthorized"}
PLATFORMS = {"youtube", "facebook", "instagram", "tiktok"}


def normalize_text(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip()).casefold()


def normalize_account_name(value: Any) -> str:
    return re.sub(r"[^a-z0-9\u4e00-\u9fff]+", "", normalize_text(value).lstrip("@"))


def normalize_profile_url(value: Any) -> str:
    raw = str(value or "").strip()
    if not raw:
        return ""
    if "://" not in raw:
        raw = "https://" + raw
    try:
        parts = urlsplit(raw)
        host = (parts.hostname or "").casefold().removeprefix("www.")
        path = re.sub(r"/+", "/", parts.path).rstrip("/").casefold()
        safe_query = urlencode(sorted((key, value) for key, value in parse_qsl(parts.query) if not key.lower().startswith("utm_")))
        return urlunsplit(("https", host, path, safe_query, ""))
    except ValueError:
        return normalize_text(raw).rstrip("/")


def split_values(value: Any) -> list[str]:
    return [item.strip() for item in re.split(r"[,，;；\n\r]+", str(value or "")) if item.strip()]


def parse_bool(value: Any) -> bool:
    return normalize_text(value) in {"1", "true", "yes", "y", "是", "允许", "可"}


def parse_date(value: Any) -> date | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip().replace("/", "-")
    for pattern in ("%Y-%m-%d", "%Y-%m", "%Y.%m.%d"):
        try:
            return datetime.strptime(text, pattern).date()
        except ValueError:
            continue
    raise ValueError(f"无法识别日期：{text}")


def _decode_csv(data: bytes) -> str:
    for encoding in ("utf-8-sig", "utf-8", "gb18030"):
        try:
            return data.decode(encoding)
        except UnicodeDecodeError:
            continue
    raise ValueError("CSV 编码无法识别，请使用 UTF-8、UTF-8 BOM 或 GB18030")


def read_sheet(filename: str, data: bytes) -> tuple[list[str], list[dict[str, Any]]]:
    suffix = filename.casefold().rsplit(".", 1)[-1]
    if suffix == "csv":
        reader = csv.DictReader(io.StringIO(_decode_csv(data)))
        headers = [str(item or "").strip() for item in (reader.fieldnames or [])]
        rows = [{str(key or "").strip(): value for key, value in row.items()} for row in reader]
        return headers, rows
    if suffix == "xlsx":
        workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        sheet = workbook.active
        iterator = sheet.iter_rows(values_only=True)
        first = next(iterator, ())
        headers = [str(item or "").strip() for item in first]
        rows = [dict(zip(headers, values)) for values in iterator if any(value not in (None, "") for value in values)]
        workbook.close()
        return headers, rows
    raise ValueError("只支持 .xlsx 和 .csv 文件")


def suggest_mapping(headers: list[str]) -> dict[str, str]:
    result: dict[str, str] = {}
    for header in headers:
        normalized = normalize_text(header)
        for field, aliases in HEADER_ALIASES.items():
            if normalized in {normalize_text(alias) for alias in aliases} and field not in result:
                result[field] = header
                break
    return result


def map_rows(headers: list[str], rows: list[dict[str, Any]], mapping: dict[str, str] | None = None) -> tuple[dict[str, str], list[dict[str, Any]]]:
    resolved = {**suggest_mapping(headers), **(mapping or {})}
    output: list[dict[str, Any]] = []
    for row_number, row in enumerate(rows, start=2):
        item = {field: row.get(header, "") for field, header in resolved.items() if header in row}
        item["_row_number"] = row_number
        output.append(item)
    return resolved, output


def validate_row(raw: dict[str, Any]) -> tuple[dict[str, Any], list[str]]:
    errors: list[str] = []
    platform = normalize_text(raw.get("platform"))
    relationship = normalize_text(raw.get("relationship_type")) or "unknown"
    authorization = normalize_text(raw.get("authorization_status")) or "unknown"
    name = str(raw.get("display_name") or "").strip()
    if platform not in PLATFORMS:
        errors.append("平台必须为 youtube/facebook/instagram/tiktok")
    if not name:
        errors.append("账号名称不能为空")
    if relationship not in RELATIONSHIPS:
        errors.append("账号类型不在标准枚举中")
    if authorization not in AUTH_STATUSES:
        errors.append("授权状态不在标准枚举中")
    try:
        start = parse_date(raw.get("authorization_start"))
        end = parse_date(raw.get("authorization_end"))
    except ValueError as exc:
        errors.append(str(exc)); start = end = None
    if start and end and start > end:
        errors.append("授权结束时间不能早于开始时间")
    return {
        "row_number": int(raw.get("_row_number") or 0),
        "platform": platform,
        "display_name": name,
        "platform_account_id": str(raw.get("platform_account_id") or "").strip(),
        "profile_url": str(raw.get("profile_url") or "").strip(),
        "relationship_type": relationship,
        "authorization_status": authorization,
        "company_name": str(raw.get("company_name") or "").strip(),
        "dramas": split_values(raw.get("dramas")),
        "authorized_regions": [item.upper() for item in split_values(raw.get("authorized_regions"))],
        "authorization_start": start,
        "authorization_end": end,
        "allowed_full_series": parse_bool(raw.get("allowed_full_series")),
        "notes": str(raw.get("notes") or "").strip(),
    }, errors


def find_account(session: Session, row: dict[str, Any]) -> MonitoredAccount | None:
    platform = row["platform"]
    if row["platform_account_id"]:
        item = session.exec(select(MonitoredAccount).where(
            MonitoredAccount.platform == platform,
            MonitoredAccount.platform_account_id == row["platform_account_id"],
        )).first()
        if item:
            return item
    normalized_url = normalize_profile_url(row["profile_url"])
    if normalized_url:
        item = session.exec(select(MonitoredAccount).where(
            MonitoredAccount.platform == platform,
            MonitoredAccount.normalized_profile_url == normalized_url,
        )).first()
        if item:
            return item
    normalized_name = normalize_account_name(row["display_name"])
    return session.exec(select(MonitoredAccount).where(
        MonitoredAccount.platform == platform,
        MonitoredAccount.normalized_name == normalized_name,
    )).first()


def drama_title_map(session: Session) -> dict[str, Drama]:
    result: dict[str, Drama] = {}
    for drama in session.exec(select(Drama)).all():
        result[normalize_text(drama.title)] = drama
    for profile in session.exec(select(DramaSearchProfile)).all():
        drama = session.get(Drama, profile.drama_id)
        if not drama:
            continue
        for name in [profile.official_title, *profile.aliases, *profile.misspellings]:
            if name:
                result[normalize_text(name)] = drama
    return result


def preview_import(filename: str, data: bytes, mapping: dict[str, str] | None = None) -> dict[str, Any]:
    headers, source_rows = read_sheet(filename, data)
    resolved, raw_rows = map_rows(headers, source_rows, mapping)
    rows: list[dict[str, Any]] = []
    invalid = 0
    for raw in raw_rows:
        item, errors = validate_row(raw)
        item["errors"] = errors
        invalid += int(bool(errors))
        rows.append(item)
    return {
        "filename": filename, "headers": headers, "fields": [{"value": field, "label": FIELD_LABELS[field]} for field in FIELDS],
        "mapping": resolved, "row_count": len(rows), "valid_count": len(rows) - invalid, "invalid_count": invalid,
        "rows": rows[:50],
    }


def import_accounts(session: Session, filename: str, data: bytes, user_id: int, mapping: dict[str, str] | None = None) -> RadarImportBatch:
    headers, source_rows = read_sheet(filename, data)
    resolved, raw_rows = map_rows(headers, source_rows, mapping)
    batch = RadarImportBatch(filename=filename, row_count=len(raw_rows), mapping_json=resolved, created_by=user_id)
    session.add(batch); session.flush()
    dramas = drama_title_map(session)
    errors: list[dict] = []
    for raw in raw_rows:
        row, row_errors = validate_row(raw)
        if row_errors:
            batch.ignored_count += 1
            errors.append({"row": row["row_number"], "errors": row_errors})
            continue
        account = find_account(session, row)
        created = account is None
        if account is None:
            account = MonitoredAccount(platform=row["platform"], display_name=row["display_name"])
        account.platform_account_id = row["platform_account_id"] or account.platform_account_id
        account.handle = row["display_name"].lstrip("@")
        account.display_name = row["display_name"]
        account.profile_url = row["profile_url"]
        account.normalized_profile_url = normalize_profile_url(row["profile_url"])
        account.normalized_name = normalize_account_name(row["display_name"])
        account.relationship_type = row["relationship_type"]
        account.authorization_status = row["authorization_status"]
        account.company_name = row["company_name"]
        account.allowed_full_series = row["allowed_full_series"]
        account.authorized_regions = row["authorized_regions"]
        account.authorization_start = row["authorization_start"]
        account.authorization_end = row["authorization_end"]
        account.notes = row["notes"]
        account.active = True
        account.source = "import"
        account.updated_at = datetime.utcnow()
        session.add(account); session.flush()
        batch.created_count += int(created); batch.updated_count += int(not created)
        for raw_title in row["dramas"]:
            drama = dramas.get(normalize_text(raw_title))
            if not drama:
                batch.unmatched_count += 1
                session.add(RadarUnmatchedDrama(import_batch_id=batch.id, row_number=row["row_number"], raw_name=raw_title))
                continue
            link = session.exec(select(MonitoredAccountDrama).where(
                MonitoredAccountDrama.account_id == account.id, MonitoredAccountDrama.drama_id == drama.id,
            )).first() or MonitoredAccountDrama(account_id=account.id, drama_id=drama.id)
            link.authorized_regions = row["authorized_regions"]
            link.authorization_start = row["authorization_start"]
            link.authorization_end = row["authorization_end"]
            link.allowed_full_series_override = row["allowed_full_series"]
            session.add(link)
    batch.errors_json = errors
    batch.status = "completed_with_errors" if errors else "completed"
    session.add(batch); session.commit(); session.refresh(batch)
    return batch


def template_bytes() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "监测账号"
    sheet.append([FIELD_LABELS[field] for field in FIELDS])
    sheet.append(["youtube", "Example Channel", "UCxxxxxxxx", "https://youtube.com/@example", "authorized_partner", "authorized", "Example Media", "Example Drama", "US,CA", "2026-01-01", "2026-12-31", "否", "示例行，导入前可删除"])
    stream = io.BytesIO(); workbook.save(stream); workbook.close()
    return stream.getvalue()


def authorization_for(
    session: Session, account: MonitoredAccount | None, drama_id: int, region: str,
    published_at: datetime | None, full_series: bool = False,
) -> dict[str, Any]:
    if not account or not account.active or account.source == "removed":
        return {"relationship_type": "unknown", "authorization_status": "unknown", "authorized": False, "reason": "未匹配监测账号"}
    link = session.exec(select(MonitoredAccountDrama).where(
        MonitoredAccountDrama.account_id == account.id, MonitoredAccountDrama.drama_id == drama_id,
    )).first()
    relationship = (link.relationship_override if link and link.relationship_override else account.relationship_type) or "unknown"
    status = (link.authorization_status_override if link and link.authorization_status_override else account.authorization_status) or "unknown"
    if account.source == "theater_match" and not link:
        return {"relationship_type": "unknown", "authorization_status": "unknown", "authorized": False, "reason": "剧场官方账号与当前剧目的剧场不匹配"}
    if relationship in {"own_official", "own_creator"}:
        return {"relationship_type": relationship, "authorization_status": "authorized", "authorized": True, "reason": "自有账号"}
    if not link:
        return {"relationship_type": relationship, "authorization_status": "unknown", "authorized": False, "reason": "未关联当前剧目"}
    start = link.authorization_start or account.authorization_start
    end = link.authorization_end or account.authorization_end
    regions = link.authorized_regions or account.authorized_regions
    allowed_full = link.allowed_full_series_override if link.allowed_full_series_override is not None else account.allowed_full_series
    day = (published_at or datetime.utcnow()).date()
    if status not in {"authorized", "partial"}:
        return {"relationship_type": relationship, "authorization_status": status, "authorized": False, "reason": "授权状态无效"}
    if start and day < start or end and day > end:
        return {"relationship_type": relationship, "authorization_status": "expired", "authorized": False, "reason": "发布时间不在授权期内"}
    if regions and region.upper() not in {item.upper() for item in regions}:
        return {"relationship_type": relationship, "authorization_status": "partial", "authorized": False, "reason": "地区不在授权范围"}
    if full_series and not allowed_full:
        return {"relationship_type": relationship, "authorization_status": "partial", "authorized": False, "reason": "仅允许推广片段"}
    return {"relationship_type": relationship, "authorization_status": status, "authorized": True, "reason": "当前剧目、地区和时间均在授权范围"}
